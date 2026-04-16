import os
import utils
import logging
from models import OnboardingFile
from sqlalchemy import update, func, cast, Date
import utils
import openpyxl
import controllers


# @description: return a list of files based on status
def returnFiles(db_conn, status="downloaded"):
    try:
        s = (
            OnboardingFile.__table__.select()
            .with_hint(OnboardingFile, "WITH (NOLOCK)")
            .where(OnboardingFile.status == status)
            .with_only_columns(
                OnboardingFile.id,
                OnboardingFile.fileName,
                OnboardingFile.providerId,
                OnboardingFile.joinDate,
                OnboardingFile.createdBy,
                OnboardingFile.productOptionId,
                OnboardingFile.brokerageId,
                OnboardingFile.orgFileName,
                OnboardingFile.scheme,
                cast(OnboardingFile.providerInceptionDate, Date).label(
                    "providerInceptionDate"
                ),
            )
        )
        # return execute query as a list of dictionaries
        result = db_conn.execute(s)
        return result.mappings().all()
    except Exception as e:
        logging.error(e)
        return []


# @description: return a list of files based on status
def returnFileById(db_conn, fileId):
    try:
        s = (
            OnboardingFile.__table__.select()
            .where(OnboardingFile.id == fileId)
            .with_only_columns(
                OnboardingFile.id,
                OnboardingFile.fileName,
                OnboardingFile.providerId,
                OnboardingFile.joinDate,
                OnboardingFile.createdBy,
                OnboardingFile.productOptionId,
                OnboardingFile.brokerageId,
                OnboardingFile.orgFileName,
                cast(OnboardingFile.providerInceptionDate, Date).label(
                    "providerInceptionDate"
                ),
            )
        )
        # return execute query as a list of dictionaries
        result = db_conn.execute(s)
        return result.mappings().all()
    except Exception as e:
        logging.error(e)
        return []


# @description: return a list of files based on status
def returnNameFile(db_conn, fileId):
    try:
        s = (
            OnboardingFile.__table__.select()
            .where(OnboardingFile.fileName == fileId)
            .with_only_columns(
                OnboardingFile.id,
                OnboardingFile.fileName,
                # OnboardingFile.productType,
                OnboardingFile.providerId,
                OnboardingFile.joinDate,
                OnboardingFile.createdBy,
                OnboardingFile.productOptionId,
                OnboardingFile.brokerageId,
                OnboardingFile.orgFileName,
                cast(OnboardingFile.providerInceptionDate, Date).label(
                    "providerInceptionDate"
                ),
            )
        )
        # return execute query as a list of dictionaries
        result = db_conn.execute(s)
        return result.mappings().all()
    except Exception as e:
        logging.error(e)
        return []


# @description: update file status
def updateFileStatus(
    db_conn, id, status, statusDescription, totalRows=0, blankRows=0, processedRows=0
):
    try:
        s = (
            update(OnboardingFile)
            .values(
                status=status,
                statusDescription=statusDescription,
                totalRows=totalRows,
                blankRows=blankRows,
                processedRows=processedRows,
            )
            .where(OnboardingFile.id == id)
        )
        db_conn.execute(s)
        db_conn.commit()
        return True
    except Exception as e:
        db_conn.rollback()
        logging.error(e)
        return False


# download files
def downloadFile():
    try:
        with utils.orm_conn(os.getenv("DATABASE_URL")) as conn:
            unprocessed = returnFiles(conn, "pending")

            if unprocessed:
                logging.debug(f"Found {len(unprocessed)} files to download")
                az_conn = utils.connectAzStorage(os.getenv("AZSAS"))
                if not az_conn:
                    logging.error("Could not connect to Azure Storage")
                    return False
                for file in unprocessed:
                    logging.debug(f"Downloading {file['fileName']}")
                    # exit()
                    # download file
                    if utils.downloadFileAzStorage(
                        az_conn,
                        os.getenv("AZ_BLOB"),
                        file["fileName"],
                        f"downloads/{file['fileName']}",
                    ):
                        logging.debug(
                            f"File {file['fileName']} downloaded successfully"
                        )

                        if not file["fileName"].endswith(".xlsx"):
                            # update status to failed
                            updateFileStatus(
                                conn,
                                file["id"],
                                "failed",
                                "File seems to be invalid",
                            )
                            utils.sendEmailNotification(
                                file["createdBy"],
                                "RMA Client Connect - File Processed",
                                f"Your file, {file["orgFileName"]} as failed to load. Please check the file and try again",
                            )
                            controllers.insertNotification(
                                conn,
                                file["createdBy"],
                                "File Processing Error",
                                f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                                "error",
                            )

                        # exit()
                        # open workbook, read only property allows for more consisttent and faster read times esecially when working on larger excel files
                        try:
                            wb = openpyxl.load_workbook(
                                "downloads/" + file["fileName"],
                                read_only=True,
                                data_only=True,
                            )
                        except Exception as e:
                            logging.error(f"Error opening workbook: {e}")
                            updateFileStatus(
                                conn,
                                file["id"],
                                "failed",
                                "File seems to be invalid",
                            )
                            utils.sendEmailNotification(
                                file["createdBy"],
                                "RMA Client Connect - File Processed",
                                f"Your file, {file["orgFileName"]} as failed to load. Please check the file and try again",
                            )
                            controllers.insertNotification(
                                conn,
                                file["createdBy"],
                                "File Processing Error",
                                f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                                "error",
                            )
                            continue
                        # set active sheet
                        sheet = wb.active

                        # check cell value
                        if sheet.cell(4, 1).value == "Company":
                            logging.debug("Company file")
                            # if cell(5,1) is blank
                            # if sheet.cell(5, 1).value == None:
                            #     updateFileStatus(
                            #         conn,
                            #         file["id"],
                            #         "downloaded",
                            #         "File has been downloaded for processing",
                            #     )

                            # check if cell(5,1) has a value and if it matches scheme
                            # elif sheet.cell(5, 1).value == file["scheme"]:
                            # logging.debug("Scheme matches")
                            # update status to downloaded
                            updateFileStatus(
                                conn,
                                file["id"],
                                "downloaded",
                                "File has been downloaded for processing",
                            )
                            # else:
                            #     updateFileStatus(
                            #         conn,
                            #         file["id"],
                            #         "failed",
                            #         "Scheme in file does not match scheme in system",
                            #     )
                        else:
                            # update status to failed
                            updateFileStatus(
                                conn,
                                file["id"],
                                "failed",
                                "File seems to be invalid",
                            )
                            utils.sendEmailNotification(
                                file["createdBy"],
                                "RMA Client Connect - File Processed",
                                f"Your file, {file["orgFileName"]} as failed to load. Please check the file and try again",
                            )
                            controllers.insertNotification(
                                conn,
                                file["createdBy"],
                                "File Processing Error",
                                f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                                "error",
                            )

                        # close workbook
                        wb.close()
    except Exception as e:
        logging.error(e)
    return False


# download files
def manualFile(fileId):
    try:
        with utils.orm_conn(os.getenv("DATABASE_URL")) as conn:
            unprocessed = returnNameFile(conn, fileId)

            if unprocessed:
                logging.debug(f"Found {len(unprocessed)} files to download")
                az_conn = utils.connectAzStorage(os.getenv("AZSAS"))
                if not az_conn:
                    logging.error("Could not connect to Azure Storage")
                    return False
                for file in unprocessed:
                    logging.debug(f"Downloading {file['fileName']}")
                    # download file
                    if utils.downloadFileAzStorage(
                        az_conn,
                        os.getenv("AZ_BLOB"),
                        file["fileName"],
                        f"downloads/{file['fileName']}",
                    ):
                        logging.debug(
                            f"File {file['fileName']} downloaded successfully"
                        )
    except Exception as e:
        logging.error(e)
    return False


def manualFileById(fileId):
    try:
        with utils.orm_conn(os.getenv("DATABASE_URL")) as conn:
            unprocessed = returnFileById(conn, fileId)

            if unprocessed:
                logging.debug(f"Found {len(unprocessed)} files to download")
                az_conn = utils.connectAzStorage(os.getenv("AZSAS"))
                if not az_conn:
                    logging.error("Could not connect to Azure Storage")
                    return False
                for file in unprocessed:
                    logging.debug(f"Downloading {file['fileName']}")
                    # download file
                    if utils.downloadFileAzStorage(
                        az_conn,
                        os.getenv("AZ_BLOB"),
                        file["fileName"],
                        f"downloads/{file['fileName']}",
                    ):
                        logging.debug(
                            f"File {file['fileName']} downloaded successfully"
                        )
    except Exception as e:
        logging.error(e)
    return False
