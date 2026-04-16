import os
import utils
import logging
from models import (
    OnboardingPolicy,
    OnboardingMember,
    OnboardingPolicyMember,
    BrokerageRepresentativeMap,
    fileDataOrg,
    onboardingPolicy,
    onboardingData,
)
from sqlalchemy import update, insert, select, and_, or_
import utils
import openpyxl
import re
import uuid
import string
from dateutil.relativedelta import *
from datetime import datetime, timedelta, date
import json
import controllers
import time


# function to format dictionary key
def fmt_key(key_name):
    # remove leading and trailling spaces
    key_name = key_name.strip()
    # replace spaces with underscores
    key_name = key_name.lower().replace(" ", "_")
    # remove non alphanumeric characters
    key_name = re.sub(r"[^a-zA-Z0-9_]", "", key_name)
    # split string on _
    key_names = key_name.split("_")
    # loop through list and capitalize each word except first
    key_name = ""
    i = 0
    for key in key_names:
        if i > 0:
            key_name += key.capitalize()
        else:
            key_name += key
        i += 1
    # return value
    return key_name


# function ro return object
def fmt_obj(field, error):
    # field = field.replace("_", " ").title()
    return {
        "field": field,
        "message": error,
    }


# check for entry on policy and fileRow
def checkPolicyEntry(db_conn, fileId, fileRow):
    try:
        s = (
            OnboardingPolicy.__table__.select()
            .distinct()
            .with_hint(OnboardingPolicy, "WITH (NOLOCK)")
            .join(
                OnboardingPolicyMember,
                OnboardingPolicy.id == OnboardingPolicyMember.policyId,
            )
            .with_hint(OnboardingPolicyMember, "WITH (NOLOCK)")
            .where(
                OnboardingPolicy.fileId == fileId,
                OnboardingPolicyMember.fileRow == fileRow,
            )
            .with_only_columns(OnboardingPolicy.id)
        )
        # print(s)
        # print(fileId)
        # print(fileRow)
        # exit()
        # return execute query as a list of dictionaries
        result = db_conn.execute(s)
        outcome = result.mappings().all()
        print(outcome)
        if outcome:
            return True
        else:
            return False
    except Exception as e:
        logging.error(e)
        exit()
    return False


def checkPolicyEntryBeneficiary(db_conn, fileId, member):
    try:
        s = (
            OnboardingPolicy.__table__.select()
            .distinct()
            .with_hint(OnboardingPolicy, "WITH (NOLOCK)")
            .join(
                OnboardingPolicyMember,
                OnboardingPolicy.id == OnboardingPolicyMember.policyId,
            )
            .with_hint(OnboardingPolicyMember, "WITH (NOLOCK)")
            .join(
                OnboardingMember, OnboardingMember.id == OnboardingPolicyMember.memberId
            )
            .with_hint(OnboardingMember, "WITH (NOLOCK)")
            .where(
                OnboardingPolicy.fileId == fileId,
                OnboardingPolicyMember.isBeneficiary == True,
                OnboardingMember.idNumber == member["idNumber"],
            )
            .with_only_columns(OnboardingPolicy.id)
        )
        # print(s)
        # print(fileId)
        # print(fileRow)
        # exit()
        # return execute query as a list of dictionaries
        result = db_conn.execute(s)
        outcome = result.mappings().all()
        print(outcome)
        if outcome:
            return True
        else:
            return False
    except Exception as e:
        logging.error(e)
        exit()
    return False


# function to create policy
def createPolicy(
    db_conn,
    joinDate,
    # productTypeId,
    providerId,
    fileId,
    coverAmount,
    createdBy,
    getPolicy=None,
    policyStatus="New",
    statusNote=None,
):
    try:
        logging.debug(f"Creating policy")
        # policyId = uuid.uuid4()

        # import pprint

        # pprint.pprint(getPolicy)
        # exit()
        if not getPolicy:
            return None

        # get brokerage representative map if exists
        s = select(BrokerageRepresentativeMap).filter(
            BrokerageRepresentativeMap.BrokerageId == getPolicy["brokerageId"],
            BrokerageRepresentativeMap.RepresentativeId
            == getPolicy["representativeId"],
        )

        result = db_conn.execute(s)
        outcome = result.mappings().all()

        brokerageMapId = outcome[0]["BrokerageRepresentativeMapId"] if outcome else None

        # import pprint

        # pprint.pprint(outcome)
        # exit()

        if not outcome:
            s = insert(BrokerageRepresentativeMap).values(
                BrokerageId=getPolicy["brokerageId"],
                RepresentativeId=getPolicy["representativeId"],
            )
            # return execute query as a list of dictionaries
            result = db_conn.execute(s)
            # db_conn.commit()
            brokerageMapId = result.inserted_primary_key[0]

        s = insert(OnboardingPolicy).values(
            joinDate=joinDate,
            # productTypeId=str(productTypeId).lower(),
            providerId=providerId,
            fileId=str(fileId).lower(),
            coverAmount=coverAmount,
            createdBy=createdBy,
            status=policyStatus,
            StatusNote=statusNote,
            AdminPercentage=getPolicy["adminPercentage"],
            CommissionPercentage=getPolicy["commissionPercentage"],
            BinderFeePercentage=getPolicy["binderFeePercentage"],
            parentPolicyId=providerId,
            SchemeRolePlayerId=getPolicy["policyOwnerId"],
            providerInceptionDate=getPolicy["policyInceptionDate"],
            productOptionId=getPolicy["productOptionId"],
            InstallmentDayOfMonth=getPolicy["regularInstallmentDayOfMonth"],
            paymentFrequencyId=getPolicy["paymentFrequencyId"],
            BrokerageRepresentativeMapId=brokerageMapId,
            ReferenceNumber=uuid.uuid4(),
        )
        # return execute query as a list of dictionaries
        result = db_conn.execute(s)
        # db_conn.commit()
        logging.debug(f"Policy created")
        # get last value insert
        return result.inserted_primary_key[0]

    except Exception as e:
        db_conn.rollback()
        logging.error(e)
        utils.sendEmailNotification(
            "clientconnect@cdasolutions.co.za",
            "RMA Client Connect - File Processed",
            f"File, {fileId} has failed to load. Please check the file and try again. {e}",
        )
        exit()
        return None


# function to update policy selected category
def updatePolicySelectedCategory(db_conn, policyId, spouse, child, extended):
    try:
        selectedCategory = 1
        if spouse or child:
            selectedCategory = 2
        if extended:
            selectedCategory = 3
        # selectedCategory = 3
        s = (
            update(OnboardingPolicy)
            .values(selectedCategory=selectedCategory)
            .where(OnboardingPolicy.id == policyId)
        )
        db_conn.execute(s)
        # db_conn.commit()
        return True
    except Exception as e:
        db_conn.rollback()
        logging.error(e)

        return False


# function to create member
def createMember(
    db_conn,
    IdTypeId,
    IdNumber,
    FirstName,
    Surname,
    DateOfBirth,
    preferredCommunicationTypeId,
    tellNumber,
    cellNumber,
    emailAddress,
    addressLine1,
    addressLine2,
    postalCode,
    city,
    province,
    countryId,
    createdBy,
    DateOfDeath=None,
):
    try:
        print("this function is called")
        s = insert(OnboardingMember).values(
            idTypeId=IdTypeId,
            idNumber=IdNumber,
            firstName=FirstName,
            surname=Surname,
            dateOfBirth=DateOfBirth,
            communicationPreferenceId=preferredCommunicationTypeId,
            telephoneNumber=tellNumber,
            mobileNumber=cellNumber,
            emailAddress=emailAddress,
            addressLine1=addressLine1,
            addressLine2=addressLine2,
            postalCode=postalCode,
            city=city,
            province=province,
            countryId=countryId,
            createdBy=createdBy,
            dateOfDeath=DateOfDeath,
        )
        # return execute query as a list of dictionaries
        res = db_conn.execute(s)
        # db_conn.commit()
        return res.inserted_primary_key[0]
    except Exception as e:
        db_conn.rollback()
        logging.error(f"error member insert {e}")
        exit()
    return None


# function to create policy member
def createPolicyMember(
    db_conn,
    policyId,
    memberId,
    createdBy,
    MemberType,
    MemberTypeId,
    Benefit,
    isBeneficiary,
    roleplayerTypeId,
    FileRow,
    exceptions,
    status,
    previousInsurer=None,
    previousInsurerPolicyNumber=None,
    previousInsurerJoinDate=None,
    previousInsurerCancellationDate=None,
    PreviousInsurerCoverAmount=0,
):
    try:
        # if memberType == "Main Member" then isBeneficiary is true
        if MemberType == "Main Member":
            isBeneficiary = True
        s = insert(OnboardingPolicyMember).values(
            policyId=policyId,
            memberId=memberId,
            createdBy=createdBy,
            memberType=MemberType,
            memberTypeId=MemberTypeId,
            benefit=Benefit,
            fileRow=FileRow,
            exceptions=exceptions,
            isBeneficiary=isBeneficiary,
            roleplayerTypeId=roleplayerTypeId,
            status=status,
            PreviousInsurer=previousInsurer,
            PreviousInsurerPolicyNumber=previousInsurerPolicyNumber,
            PreviousInsurerJoinDate=previousInsurerJoinDate,
            PreviousInsurerCancellationDate=previousInsurerCancellationDate,
            PreviousInsurerCoverAmount=PreviousInsurerCoverAmount,
        )
        # return execute query as a list of dictionaries
        db_conn.execute(s)
        # db_conn.commit()
        return True
    except Exception as e:
        db_conn.rollback()
        logging.error(e)
        utils.sendEmailNotification(
            "clientconnect@cdasolutions.co.za",
            "RMA Client Connect - File Processed",
            f"FileRow, {FileRow} has failed to load. Please check the file and try again. {e}",
        )
        exit()
        return False


# update as beneficiary or create new beneficiary
def updateAsBeneficiary(db_conn, policyId, member):
    setBeneficiary = False
    try:

        # check if beneficiary exists
        s = (
            OnboardingPolicyMember.__table__.select()
            .join(
                OnboardingMember, OnboardingMember.id == OnboardingPolicyMember.memberId
            )
            .where(
                OnboardingPolicyMember.policyId == policyId,
                OnboardingMember.idNumber == member["idNumber"],
            )
            .with_only_columns(
                OnboardingPolicyMember.policyId,
                OnboardingPolicyMember.memberId,
                OnboardingPolicyMember.isBeneficiary,
                OnboardingPolicyMember.memberType,
            )
        )
        # return execute query as a list of dictionaries
        result = db_conn.execute(s)
        outcome = result.mappings().all()
        for row in outcome:
            # logging.debug(row)
            if row["isBeneficiary"] == False:
                s = (
                    update(OnboardingPolicyMember)
                    .values(isBeneficiary=True)
                    .where(
                        OnboardingPolicyMember.policyId == policyId,
                        OnboardingPolicyMember.memberId == row["InsuredMemberId"],
                    )
                )
                db_conn.execute(s)
                # db_conn.commit()
                setBeneficiary = True

        # if not setBeneficiary:
        #     memberId = createMember(
        #         db_conn,
        #         member["idTypeId"],
        #         member["idNumber"] if member["idNumber"] else member["passportNumber"],
        #         member["firstName"],
        #         member["surname"],
        #         member["dateOfBirth"],
        #         utils.returnCommunicationType(member["preferredMethodOfCommunication"]),
        #         member["telephone"],
        #         member["mobile"],
        #         member["email"],
        #         member["address1"],
        #         member["address2"],
        #         member["postalCode"],
        #         member["city"],
        #         member["province"],
        #         1,
        #         member["createdBy"],
        #     )

        #     logging.debug(f"Member created {memberId}")

        #     createPolicyMember(
        #         db_conn,
        #         policyId,
        #         memberId,
        #         member["createdBy"],
        #         member["clientType"],
        #         member["memberTypeId"],
        #         member["benefitName"],
        #         member["isBeneficiary"],
        #         member["roleplayerTypeId"],
        #         member["fileRow"],
        #         member["exceptions"],
        #         member["status"],
        #     )

        #     logging.debug(f"Member created {memberId}")
        #     db_conn.commit()

        # db_conn.commit()

    except Exception as e:
        db_conn.rollback()
        logging.error(e)
        # exit()
    return setBeneficiary


# process the excel file
def processFiles():
    with utils.orm_conn(os.getenv("DATABASE_URL")) as conn:
        downloaded = controllers.returnFiles(conn)

        for file in downloaded:
            try:

                # set rows
                totalRows = 0
                blankRows = 0
                processedRows = 0
                hasMainMember = False

                logging.debug(f"Processing {file['fileName']}")
                # open workbook, read only property allows for more consisttent and faster read times esecially when working on larger excel files
                wb = openpyxl.load_workbook(
                    "downloads/" + file["fileName"], read_only=True, data_only=True
                )

                logging.debug(f"Processing {file['fileName']}")

                # set active sheet
                sheet = wb.active

                # get list of column names
                key_names = []
                for x in range(1, sheet.max_column + 1):
                    if sheet.cell(4, x).value:
                        key_names.append(
                            fmt_key(sheet.cell(4, x).value)
                        )  # column names start at row 4

                    # if column is blank then check if rows below are blank then said column name as blank + x
                    else:
                        key_names.append(fmt_key(f"blank{x}"))

                row_values = []
                policyId = None
                spouse = False
                child = False
                extended = False
                getPolicy = None
                spouseCnt = 0
                childCnt = 0
                beneficiaries = []

                row_num = 5  # rows starts at row 5
                mx_row = sheet.max_row  # last row in sheet
                totalRows = mx_row - row_num + 1

                # convert rows to dictionary
                # iterate over rows
                for row in sheet.iter_rows(min_row=row_num, values_only=True):
                    member = []
                    policyMember = []
                    exceptions = []

                    logging.debug(f"processing {row_num} of {mx_row}")
                    # set dictinary key:value for each column
                    member = {key_names[i]: row[i] for i in range(len(key_names))}
                    hasBlank = True

                    # check if blank row based on these columns being blank
                    for key, value in member.items():
                        if key in [
                            "clientType",
                            "firstName",
                            "surname",
                            "idNumber",
                            "passportNumber",
                        ]:
                            if value:
                                hasBlank = False
                                break

                    # member row
                    member["fileRow"] = row_num
                    row_num += 1
                    if hasBlank:
                        blankRows += 1
                        continue

                    # check if file has been processed previously file["id"]
                    if checkPolicyEntry(conn, file["id"], member["fileRow"]):
                        processedRows += 1
                        # utils.sendEmailNotification("lourens@cdasolutions.co.za", "RMA Client Connect - File Already Processed", f"Your file, {file["orgFileName"]} has failed to load. Possible duplicates")
                        # conn.rollback()
                        # exit()
                        continue

                    processedRows += 1

                    # firstname START
                    # not blank and should be upper case and match to VOPD info
                    member["firstName"] = (
                        str(member["firstName"]).strip().upper().replace(",", "")
                        if member["firstName"]
                        else None
                    )
                    # check if value entered
                    if member["firstName"] == None:
                        exceptions.append(
                            fmt_obj("firstName", "First name cannot be blank")
                        )
                    # check if value contains a number
                    elif utils.contains_number(member["firstName"]):
                        exceptions.append(
                            fmt_obj(
                                "firstName",
                                "First name cannot contain a number",
                            )
                        )
                    # firstname END

                    # surname START
                    # not blank and should be upper case and match to VOPD info
                    member["surname"] = (
                        str(member["surname"]).strip().upper().replace(",", "")
                        if member["surname"]
                        else None
                    )
                    # check if value entered
                    if member["surname"] == None:
                        exceptions.append(fmt_obj("surname", "Surname cannot be blank"))
                    # check if value contains a number
                    elif utils.contains_number(member["surname"]):
                        exceptions.append(
                            fmt_obj("surname", "Surname cannot contain a number")
                        )
                    # surname END

                    # clientType START
                    member["clientType"] = (
                        str(member["clientType"]).lower().strip()
                        if member["clientType"]
                        else None
                    )

                    # standardise clientType

                    member["memberTypeId"] = 0
                    if member["clientType"]:
                        if member["clientType"].lower() in [
                            "extend",
                            "extended",
                            "family member",
                            "familymember",
                        ]:
                            member["clientType"] = "Extended Family"
                        member["clientType"] = string.capwords(member["clientType"])
                        if member["clientType"] not in [
                            "Main Member",
                            "Spouse",
                            "Family Member",
                            "Child",
                            "Extended Family",
                            "Beneficiary",
                        ] and "main" in member["clientType"].lower().replace(" ", ""):
                            member["clientType"] = "Main Member"
                        if member["clientType"] not in [
                            "Main Member",
                            "Spouse",
                            "Family Member",
                            "Child",
                            "Extended Family",
                            "Beneficiary",
                        ]:
                            exceptions.append(
                                fmt_obj(
                                    "MemberType",
                                    "Invalid member type specified.",
                                )
                            )
                            member["memberTypeId"] = 0
                        else:
                            member["memberTypeId"] = utils.returnMemberType(
                                member["clientType"]
                            )
                    elif not member["clientType"]:
                        exceptions.append(
                            fmt_obj(
                                "clientType",
                                "Invalid client type specified.",
                            )
                        )

                    member["roleplayerTypeId"] = utils.returnRolePlayerType(
                        member["memberTypeId"]
                    )
                    member["isBeneficiary"] = (
                        True if member["clientType"] == "Beneficiary" else False
                    )

                    if member["clientType"] == "Spouse":
                        spouse = True
                        spouseCnt += 1
                        if spouseCnt > 1:
                            exceptions.append(
                                fmt_obj(
                                    "clientType",
                                    "Only one spouse is allowed per policy",
                                )
                            )

                    if member["clientType"] == "Child":
                        child = True
                        childCnt += 1
                        if childCnt > 6:
                            exceptions.append(
                                fmt_obj(
                                    "clientType",
                                    "Only six children are allowed per policy",
                                )
                            )
                    if member["clientType"] == "Extended Family":
                        extended = True
                    # clientType END

                    # idNumber START
                    member["idNumber"] = (
                        str(member["idNumber"])
                        .strip()
                        .replace(".", "")
                        .replace("'", "")
                        .replace(" ", "")
                        if member["idNumber"]
                        else None
                    )

                    member["passportNumber"] = (
                        str(member["passportNumber"]).strip().replace(" ", "")
                        if member["passportNumber"]
                        else None
                    )

                    # add leading zeros on child client types
                    if (
                        # member["clientType"] == "Child"
                        # and
                        member["idNumber"]
                        and len(member["idNumber"]) in [11, 12]
                    ):
                        member["idNumber"] = ("00" + member["idNumber"])[-13:]

                    if (
                        member["idNumber"] == None
                        and "main" in member["clientType"].lower()
                        and member["mainMemberLinkId"]
                    ):
                        member["idNumber"] = member["mainMemberLinkId"]

                    # check if valid SA id
                    valid_id = utils.validate_idno(member["idNumber"])
                    if (
                        not valid_id
                        and member["idNumber"]
                        and member["passportNumber"] == None
                        and "main" in member["clientType"].lower()
                    ):
                        exceptions.append(
                            fmt_obj(
                                "idNumber",
                                "Invalid ID number supplied",
                            )
                        )
                        member["idTypeId"] = 1
                    else:
                        member["idTypeId"] = 1 if valid_id else 2

                    if (
                        not (member["idNumber"])
                        and not (member["passportNumber"])
                        and member["clientType"]
                        in [
                            "Main Member",
                            # "Spouse",
                        ]
                    ):
                        exceptions.append(
                            fmt_obj(
                                "idNumber",
                                "No ID number or passport supplied",
                            )
                        )

                    # idNumber END

                    # dob check START
                    # if valid id calculate DOB
                    member["dateOfBirth"] = (
                        utils.return_dob(member["idNumber"])
                        if valid_id
                        else member["dateOfBirth"]
                    )

                    try:

                        if type(member["dateOfBirth"]) == str:
                            check_date = datetime.strptime(
                                member["dateOfBirth"], "%Y-%m-%d"
                            )
                        else:
                            check_date = member["dateOfBirth"]
                        member["dateOfBirth"] = check_date.strftime("%Y-%m-%d")
                    except:
                        member["dateOfBirth"] = None

                    if not (member["dateOfBirth"]):
                        try:
                            if member["idNumber"] and "-" in member["idNumber"]:
                                check_date = datetime.strptime(
                                    member["idNumber"], "%d-%m-%y"
                                )
                            elif member["idNumber"] and "/" in member["idNumber"]:
                                check_date = datetime.strptime(
                                    member["idNumber"], "%d/%m/%y"
                                )
                            elif (
                                not (member["idNumber"])
                                and member["passportNumber"]
                                and "-" in member["passportNumber"]
                            ):
                                check_date = datetime.strptime(
                                    member["passportNumber"], "%d-%m-%Y"
                                )
                            elif (
                                not (member["idNumber"])
                                and member["passportNumber"]
                                and "/" in member["passportNumber"]
                            ):
                                check_date = datetime.strptime(
                                    member["passportNumber"], "%d/%m/%Y"
                                )
                            elif not (member["idNumber"]) and member["passportNumber"]:
                                check_date = datetime.strptime(
                                    member["passportNumber"], "%Y-%m-%d %H:%M:%S"
                                )
                            else:
                                check_date = datetime.strptime(
                                    member["idNumber"][:6], "%y%m%d"
                                )

                            member["dateOfBirth"] = check_date.strftime("%Y-%m-%d")

                        except:
                            member["dateOfBirth"] = None

                    if not (member["dateOfBirth"]):
                        try:
                            if member["idNumber"] and "-" in member["idNumber"]:
                                check_date = datetime.strptime(
                                    member["idNumber"], "%Y-%m-%d"
                                )
                            elif member["idNumber"] and "/" in member["idNumber"]:
                                check_date = datetime.strptime(
                                    member["idNumber"], "%Y/%m/%d"
                                )
                            elif (
                                not (member["idNumber"])
                                and member["passportNumber"]
                                and "-" in member["passportNumber"]
                            ):
                                check_date = datetime.strptime(
                                    member["passportNumber"], "%Y-%m-%d"
                                )

                            elif (
                                not (member["idNumber"])
                                and member["passportNumber"]
                                and "/" in member["passportNumber"]
                            ):
                                check_date = datetime.strptime(
                                    member["passportNumber"], "%Y/%m/%d"
                                )

                            member["dateOfBirth"] = check_date.strftime("%Y-%m-%d")
                        except:
                            member["dateOfBirth"] = None

                    # correction for date of birth being in the future
                    if (
                        member["dateOfBirth"]
                        and datetime.strptime(member["dateOfBirth"], "%Y-%m-%d")
                        > datetime.now()
                    ):
                        member["dateOfBirth"] = (
                            datetime.strptime(member["dateOfBirth"], "%Y-%m-%d")
                            - relativedelta(years=100)
                        ).strftime("%Y-%m-%d")

                    if not (member["dateOfBirth"]):

                        exceptions.append(
                            fmt_obj(
                                "dateOfBirth",
                                "No valid DOB supplied",
                            )
                        )
                    # dob check END

                    # check age
                    if member["dateOfBirth"]:
                        age = utils.return_age(
                            datetime.strptime(member["dateOfBirth"], "%Y-%m-%d"),
                            file["joinDate"],
                        )
                        if member["clientType"] == "Main Member" and age < 18:
                            exceptions.append(
                                fmt_obj(
                                    "dateOfBirth",
                                    "Main member must be 18 years or older",
                                )
                            )

                        # check dob child
                        if member["clientType"] == "Child" and age > 21:
                            tempDOB = datetime.strptime(
                                member["dateOfBirth"], "%Y-%m-%d"
                            ) + relativedelta(years=100)

                            age = utils.return_age(
                                tempDOB,
                                file["joinDate"],
                            )

                            if age > 25:
                                exceptions.append(
                                    fmt_obj(
                                        "dateOfBirth",
                                        "Child must be younger than 25 years",
                                    )
                                )
                            else:
                                member["dateOfBirth"] = tempDOB.strftime("%Y-%m-%d")

                    # check VOPD for ID number START
                    # if valid id check VOPD
                    member["status"] = "New"

                    if valid_id:
                        # check for similary update date of birth and name
                        getMember = controllers.returnVOPD(conn, member["idNumber"])

                        if getMember:
                            # update name and surname if similarity is greater than 50%
                            firstNameSimilarity = utils.similarityScores(
                                getMember[0]["firstName"], member["firstName"]
                            )
                            if (
                                firstNameSimilarity["percentage"] > 0
                                or firstNameSimilarity["levenshtein"] >= 0.5
                            ):
                                member["firstName"] = getMember[0]["firstName"]
                            else:
                                exceptions.append(
                                    fmt_obj(
                                        "firstName",
                                        f"First name does not match VOPD - {getMember[0]['firstName']}",
                                    )
                                )

                            surnameSimilarity = utils.similarityScores(
                                getMember[0]["surname"], member["surname"]
                            )

                            if (
                                surnameSimilarity["percentage"] > 0
                                or surnameSimilarity["levenshtein"] >= 0.5
                            ):
                                member["surname"] = getMember[0]["surname"]
                            else:
                                exceptions.append(
                                    fmt_obj(
                                        "surname",
                                        f"Surname does not match VOPD {getMember[0]['surname']}",
                                    )
                                )

                            member["dateOfBirth"] = getMember[0]["dateOfBirth"]
                            # if person deceased then rerunning vopd doesn't matter
                            if getMember[0]["dateOfDeath"]:
                                member["dateOfDeath"] = getMember[0]["dateOfDeath"]
                                exceptions.append(
                                    fmt_obj(
                                        "idNumber",
                                        "Member is deceased",
                                    )
                                )
                                member["vopdVerified"] = True
                                member["vopdVerificationDate"] = getMember[0][
                                    "updatedAt"
                                ]
                                member["status"] = "Error"
                            # if checked in the last 30 days
                            elif getMember[0]["updatedAt"] >= (
                                datetime.now() - timedelta(days=30)
                            ):
                                member["vopdVerified"] = True
                                member["vopdVerificationDate"] = getMember[0][
                                    "updatedAt"
                                ]
                                member["status"] = "VOPD Complete"
                        # elif member["clientType"] == "Main Member":
                        else:
                            if controllers.addVOPD(conn, member["idNumber"]):
                                member["status"] = "Waiting for VOPD"
                    # check VOPD for ID number END

                    # email check START
                    member["email"] = (
                        str(member["email"]).strip() if member["email"] else None
                    )
                    if member["email"] and not (utils.valid_email(member["email"])):
                        exceptions.append(
                            fmt_obj(
                                "email",
                                "Invalid e-mail address supplied, check format.",
                            )
                        )
                    # email check END

                    # preferredMethodOfCommunication check START
                    member["preferredMethodOfCommunication"] = (
                        str(member["preferredMethodOfCommunication"])
                        .strip()
                        .replace("-", "")
                        if member["preferredMethodOfCommunication"]
                        else None
                    )

                    if member["preferredMethodOfCommunication"]:
                        if str(member["preferredMethodOfCommunication"]).lower() in [
                            "phone",
                            "post",
                            "email",
                        ]:
                            member["preferredMethodOfCommunication"] = member[
                                "preferredMethodOfCommunication"
                            ].capitalize()

                        elif member["preferredMethodOfCommunication"].lower() == "sms":
                            member["preferredMethodOfCommunication"] = member[
                                "preferredMethodOfCommunication"
                            ].upper()

                        else:
                            exceptions.append(
                                fmt_obj(
                                    "preferredMethodOfCommunication",
                                    "Invalid preference of communication",
                                )
                            )

                    if not member["preferredMethodOfCommunication"]:
                        if member["email"]:
                            member["preferredMethodOfCommunication"] = "Email"
                        elif member["mobile"]:
                            member["preferredMethodOfCommunication"] = "SMS"
                        elif member["telephone"]:
                            member["preferredMethodOfCommunication"] = "Phone"

                    if (
                        not member["preferredMethodOfCommunication"]
                        and member["clientType"] == "Main Member"
                    ):
                        exceptions.append(
                            fmt_obj(
                                "preferredMethodOfCommunication",
                                "No preferred method of communication specified",
                            )
                        )
                    # preferredMethodOfCommunication check END

                    member["address1"] = (
                        str(member["address1"]).upper().strip()
                        if member["address1"]
                        else None
                    )
                    member["address2"] = (
                        str(member["address2"]).upper().strip()
                        if member["address2"]
                        else None
                    )
                    member["city"] = (
                        str(member["city"]).upper().strip() if member["city"] else None
                    )
                    member["province"] = (
                        str(member["province"]).upper().strip()
                        if member["province"]
                        else None
                    )
                    member["country"] = (
                        str(member["country"]).upper().strip()
                        if member["country"]
                        else None
                    )

                    member["areaCode"] = (
                        str(member["areaCode"]).strip() if member["areaCode"] else None
                    )

                    if (
                        member["address1"]
                        and member["address2"]
                        and member["address2"] in member["address1"]
                    ):
                        # remove address2 from address1 and trim
                        member["address1"] = (
                            member["address1"].replace(member["address2"], "").strip()
                        )

                    # remove city from address1 and trim
                    if (
                        member["address1"]
                        and member["city"]
                        and member["city"]
                        and member["city"] in member["address1"]
                    ):
                        member["address1"] = (
                            member["address1"].replace(member["city"], "").strip()
                        )

                    # if province and address 1 then check if province is in address 1 and remove and trim
                    if (
                        member["province"]
                        and member["address1"]
                        and member["province"] in member["address1"]
                    ):
                        member["address1"] = (
                            member["address1"].replace(member["province"], "").strip()
                        )

                    # if postal code and address 1 then check if postal code is in address 1 and remove and trim
                    if (
                        member["areaCode"]
                        and member["address1"]
                        and member["areaCode"] in member["address1"]
                    ):
                        member["address1"] = (
                            member["address1"].replace(member["areaCode"], "").strip()
                        )

                    if not member["address1"] and member["clientType"] == "Main Member":
                        exceptions.append(
                            fmt_obj(
                                "address1",
                                "No address details specified",
                            )
                        )

                    # if member["address1"] and address1 is greater than 50 characters then error
                    if member["address1"] and len(member["address1"]) > 50:
                        exceptions.append(
                            fmt_obj(
                                "address1",
                                "Address 1 cannot be longer than 50 characters",
                            )
                        )

                    # if member["address2"] and address1 is greater than 50 characters then error
                    if member["address2"] and len(member["address2"]) > 50:
                        exceptions.append(
                            fmt_obj(
                                "address2",
                                "Address 2 cannot be longer than 50 characters",
                            )
                        )

                    # remove all non numbers from areaCode
                    if member["areaCode"]:
                        member["areaCode"] = re.sub(r"[^0-9]", "", member["areaCode"])

                    # if areaCode is specified and longer than 6 characters then clear
                    if member["areaCode"] and len(member["areaCode"]) > 6:
                        member["areaCode"] = None

                    # if areaCode is specified and longer than 4 characters then error
                    if member["areaCode"] and len(member["areaCode"]) > 4:
                        exceptions.append(
                            fmt_obj(
                                "areaCode",
                                "Area code cannot be longer than 4 characters",
                            )
                        )

                    # if address1 is specified then must have postal code
                    # if member["address1"] and not member["areaCode"]:
                    #     exceptions.append(
                    #         fmt_obj(
                    #             "postalCode",
                    #             "Postal code required",
                    #         )
                    #     )

                    # if address1 is specified then must have province
                    # if member["address1"] and not member["province"]:
                    #     exceptions.append(
                    #         fmt_obj(
                    #             "province",
                    #             "Province required",
                    #         )
                    #     )

                    list_of_provinces = [
                        "EASTERN CAPE",
                        "FREE STATE",
                        "GAUTENG",
                        "KWAZULU-NATAL",
                        "LIMPOPO",
                        "MPUMALANGA",
                        "NORTH WEST",
                        "NORTHERN CAPE",
                        "WESTERN CAPE",
                    ]

                    if member["province"] and "KWA" in member["province"]:
                        member["province"] = "KWAZULU-NATAL"

                    if (
                        member["province"]
                        and member["province"] not in list_of_provinces
                    ):
                        exceptions.append(
                            fmt_obj(
                                "province",
                                f"Invalid province specified - {member["province"]}",
                            )
                        )
                        member["province"] = None

                    # member["postalAddress1"] = (
                    #     str(member["postalAddress1"]).upper().strip()
                    #     if member["postalAddress1"]
                    #     else None
                    # )
                    # member["postalAddress2"] = (
                    #     str(member["postalAddress2"]).upper().strip()
                    #     if member["postalAddress2"]
                    #     else None
                    # )
                    # member["postalCity"] = (
                    #     str(member["postalCity"]).upper().strip()
                    #     if member["postalCity"]
                    #     else None
                    # )
                    # member["postalProvince"] = (
                    #     str(member["postal_province"]).upper().strip()
                    #     if member["postal_province"]
                    #     else None
                    # )
                    # member["postal_country"] = (
                    #     str(member["postal_country"]).upper().strip()
                    #     if member["postal_country"]
                    #     else None
                    # )
                    # logging.debug(member)

                    # previous insurance check START

                    member["previousInsurerPolicyNumber"] = (
                        str(member["previousInsurerPolicyNumber"]).upper().strip()
                        if member["previousInsurerPolicyNumber"]
                        else None
                    )

                    member["previousInsurerJoinDate"] = (
                        member["previousInsurerJoinDate"]
                        if member["previousInsurerJoinDate"]
                        and type(member["previousInsurerJoinDate"]) == datetime
                        else None
                    )

                    member["previousInsurerCancellationDate"] = (
                        member["previousInsurerCancellationDate"]
                        if member["previousInsurerCancellationDate"]
                        and type(member["previousInsurerJoinDate"]) == datetime
                        else None
                    )

                    member["PreviousInsurerCoverAmount"] = 0

                    # check if beneficiary already flagged
                    if checkPolicyEntryBeneficiary(conn, file["id"], member):
                        continue

                    if member["clientType"] == "Beneficiary" and policyId:
                        resultSetBeneficiary = updateAsBeneficiary(
                            conn, policyId, member
                        )
                        if resultSetBeneficiary:
                            continue

                    if member["clientType"] == "Main Member" or not policyId:
                        if member["clientType"] == "Main Member":
                            hasMainMember = True
                        # if (
                        #     member["clientType"] == "Main Member"
                        #     and policyId
                        #     and beneficiaries
                        # ):
                        #     for beneficiary in beneficiaries:
                        #         logging.debug(beneficiary)
                        #         updateAsBeneficiary(conn, policyId, beneficiary)
                        #     beneficiaries = []
                        if policyId:
                            updatePolicySelectedCategory(
                                conn, policyId, spouse, child, extended
                            )
                        logging.debug(member["benefitName"])
                        coverAmount = utils.returnCoverAmount(member["benefitName"])
                        if member["clientType"] != "Beneficiary":
                            # get cover amount from policy
                            if coverAmount == 0:
                                benefits = controllers.getMemberBenefitByName(
                                    conn, member["benefitName"]
                                )
                                if benefits:
                                    coverAmount = benefits["benefitAmount"]

                            # if coverAmount == 0:
                            #     utils.sendEmailNotification(f"lourens@cdasolutions.co.za", "RMA Client Connect - File issue benefit 0", f"Your file, {file['orgFileName']} has failed to load. Please check the file and try again")
                            #     conn.rollback()
                            #     exit()

                        policyStatus = "Processing"
                        if len(exceptions) > 0:
                            policyStatus = "Error"

                        if getPolicy == None:
                            getPolicy = utils.get_rma_api(
                                f"/clc/api/Policy/Policy/{file['providerId']}"
                            )
                        if getPolicy == None:
                            wb.close()
                            logging.error("Error getting policy")
                            utils.sendEmailNotification(
                                f"lourens@cdasolutions.co.za",
                                "RMA Client Connect - File issue benefit 0",
                                f"Your file, {file['orgFileName']} has failed to load. Can't get policy",
                            )
                            conn.rollback()
                            # time.sleep(30)
                            exit()
                            continue

                        # productOptionId = getPolicy["productOptionId"]

                        # # add this check back later LOURENS 20240417
                        # # get benefits for policy /clc/api/Product/Benefit/GetProductBenefitRates/{productOptionId}/1
                        # getBenefits = utils.get_rma_api(
                        #     f"/clc/api/Product/Benefit/GetProductBenefitRates/{productOptionId}/1"
                        # )

                        # if coverAmount == 0 and getBenefits and "benefits" in getBenefits and member["benefitName"]:
                        #     benefitExists = False
                        #     for benefit in getBenefits["benefits"]:
                        #         if benefit["name"].replace(" ", "").lower() == member["benefitName"].replace(" ", "").lower():
                        #             coverAmount = benefit["benefitRate"][0]["benefitAmount"] if "benefitRate" in benefit and benefit["benefitRate"] else 0
                        #             benefitExists = True
                        #             break

                        statusNote = None

                        # if coverAmount == 0 and not benefitExists:
                        #     coverAmount = 0

                        # if coverAmount == 0:
                        #     policyStatus = "Error"
                        #     statusNote = "Could not allocate cover amount to policy"

                        if coverAmount == 0:
                            policyStatus = "Error"
                            statusNote = "Could not allocate cover amount to policy"

                        policyId = createPolicy(
                            conn,
                            file["joinDate"],
                            # file["productTypeId"],
                            file["providerId"],
                            file["id"],
                            coverAmount,
                            file["createdBy"],
                            getPolicy,
                            policyStatus,
                            statusNote,
                        )
                        spouse = False
                        child = False
                        extended = False
                        childCnt = 0
                        spouseCnt = 0

                        if not policyId:
                            conn.rollback()
                            controllers.updateFileStatus(
                                conn, file["id"], "Error", "Error creating policy"
                            )
                            utils.sendEmailNotification(
                                file["createdBy"],
                                "RMA Client Connect - File Processed",
                                f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                            )
                            controllers.insertNotification(
                                conn,
                                file["createdBy"],
                                "File Processing Error",
                                f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                                "error",
                            )
                            break

                    if len(exceptions) > 0:
                        member["status"] = "Error"
                        if controllers.updatePolicyStatus(conn, policyId, "Error"):
                            logging.debug("Updated policy status to error")

                    # if member["clientType"] == "Beneficiary":
                    #     member["createdBy"] = file["createdBy"]
                    #     member["rowNum"] = member["fileRow"]
                    #     member["exceptions"] = json.dumps(exceptions)
                    #     member["policyId"] = policyId
                    #     beneficiaries.append(member)
                    #     logging.debug(f"Beneficiary queue for end of file")
                    #     continue
                    # logging.debug(f"Member {member}")
                    memberId = createMember(
                        conn,
                        member["idTypeId"],
                        (
                            member["idNumber"]
                            if member["idNumber"]
                            else member["passportNumber"]
                        ),
                        member["firstName"],
                        member["surname"],
                        member["dateOfBirth"],
                        utils.returnCommunicationType(
                            member["preferredMethodOfCommunication"]
                        ),
                        member["telephone"],
                        member["mobile"],
                        member["email"],
                        member["address1"],
                        member["address2"],
                        member["areaCode"],
                        member["city"],
                        member["province"],
                        1,
                        file["createdBy"],
                        DateOfDeath=member["dateOfDeath"],
                    )

                    logging.debug(f"Member created {memberId}")

                    createPolicyMember(
                        conn,
                        policyId,
                        memberId,
                        file["createdBy"],
                        member["clientType"],
                        member["memberTypeId"],
                        member["benefitName"],
                        member["isBeneficiary"],
                        member["roleplayerTypeId"],
                        member["fileRow"],
                        json.dumps(exceptions),
                        member["status"],
                        member["previousInsurer"],
                        member["previousInsurerPolicyNumber"],
                        member["previousInsurerJoinDate"],
                        member["previousInsurerCancellationDate"],
                    )

                if policyId:
                    updatePolicySelectedCategory(
                        conn, policyId, spouse, child, extended
                    )

                logging.debug(f"hasMainMember {hasMainMember}")
                if hasMainMember:
                    conn.commit()
                    controllers.updateFileStatus(
                        conn,
                        file["id"],
                        "Uploaded",
                        "Uploaded members",
                        totalRows,
                        blankRows,
                        processedRows,
                    )

                    utils.sendEmailNotification(
                        file["createdBy"],
                        "RMA Client Connect - File Processed",
                        f"Your file, {file["orgFileName"]} has been loaded successfully and is being processed.",
                        contentType="HTML",
                    )
                    controllers.insertNotification(
                        conn,
                        file["createdBy"],
                        "File Processed",
                        f"Your file, {file["orgFileName"]} has been loaded successfully and is being processed.",
                        link=f"{os.getenv("APP_LINK")}/MyPolicies?fileId={file["id"]}",
                    )
                else:
                    logging.error("No main member")
                    conn.rollback()
                    controllers.updateFileStatus(
                        conn,
                        file["id"],
                        "Error",
                        "Please check file - Member type issue",
                        totalRows,
                    )
                    utils.sendEmailNotification(
                        file["createdBy"],
                        "RMA Client Connect - File Processed",
                        f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                    )
                    controllers.insertNotification(
                        conn,
                        file["createdBy"],
                        "File Processing Error",
                        f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                        "error",
                    )

                wb.close()

                # if beneficiaries:
                #     for beneficiary in beneficiaries:
                #         logging.debug(beneficiary)
                #         updateAsBeneficiary(conn, beneficiary["policyId"], beneficiary)

                # exit()

                # os.remove("downloads/" + file["fileName"])

            except Exception as e:
                logging.error(e)
                conn.rollback()
                controllers.updateFileStatus(
                    conn,
                    file["id"],
                    "Error",
                    "Please check the file and try again",
                    totalRows,
                )
                utils.sendEmailNotification(
                    file["createdBy"],
                    "RMA Client Connect - File Processed",
                    f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                )
                controllers.insertNotification(
                    conn,
                    file["createdBy"],
                    "File Processing Error",
                    f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                    "error",
                )
                utils.sendEmailNotification(
                    "clientconnect@cdasolutions.co.za",
                    "RMA Client Connect - File Processed",
                    f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again {e}",
                )
                exit()


# insert data into fileDataOrg
def insertFileDataOrg(conn, data):
    try:
        # check if fileDataOrg entry exists for fileId and fileRow
        s = (
            select(fileDataOrg)
            .with_hint(fileDataOrg, "WITH (NOLOCK)")
            .where(
                and_(
                    fileDataOrg.fileId == data["fileId"],
                    fileDataOrg.fileRow == data["fileRow"],
                )
            )
            .with_only_columns(fileDataOrg.fileId, fileDataOrg.fileRow)
        )
        rows = conn.execute(s)
        result = rows.mappings().all()
        if result:
            logging.debug(
                f"FileDataOrg entry exists for {data['fileId']} - {data['fileRow']}"
            )
            return

        s = insert(fileDataOrg).values(data)
        conn.execute(s)
    except Exception as e:
        logging.error(e)


# process the excel file for original data
def processFilesOrg():
    with utils.orm_conn(os.getenv("DATABASE_URL")) as conn:
        downloaded = controllers.returnFiles(conn)

        for file in downloaded:
            try:
                fileId = file["id"]

                # set rows
                totalRows = 0
                blankRows = 0
                processedRows = 0
                hasMainMember = False

                logging.debug(f"Processing {file['fileName']}")
                # open workbook, read only property allows for more consisttent and faster read times esecially when working on larger excel files
                wb = openpyxl.load_workbook(
                    "downloads/" + file["fileName"], read_only=True, data_only=True
                )

                logging.debug(f"Processing {file['fileName']}")

                # set active sheet
                sheet = wb.active

                # get list of column names
                key_names = []
                for x in range(1, sheet.max_column + 1):
                    if sheet.cell(4, x).value:
                        key_names.append(
                            fmt_key(sheet.cell(4, x).value)
                        )  # column names start at row 4

                    # if column is blank then check if rows below are blank then said column name as blank + x
                    else:
                        key_names.append(fmt_key(f"blank{x}"))

                row_values = []
                policyId = None
                spouse = False
                child = False
                extended = False
                getPolicy = None
                spouseCnt = 0
                childCnt = 0
                beneficiaries = []

                row_num = 5  # rows starts at row 5
                mx_row = sheet.max_row  # last row in sheet
                totalRows = mx_row - row_num + 1

                # convert rows to dictionary
                # iterate over rows
                for row in sheet.iter_rows(min_row=row_num, values_only=True):
                    member = []
                    policyMember = []
                    exceptions = []

                    logging.debug(f"processing {row_num} of {mx_row}")
                    # set dictinary key:value for each column
                    member = {key_names[i]: row[i] for i in range(len(key_names))}
                    hasBlank = True

                    # check if blank row based on these columns being blank
                    for key, value in member.items():
                        if key in [
                            "clientType",
                            "firstName",
                            "surname",
                            "idNumber",
                            "passportNumber",
                        ]:
                            if value:
                                hasBlank = False
                                break

                    # member row
                    member["fileRow"] = row_num
                    member["fileId"] = fileId
                    row_num += 1
                    if hasBlank:
                        blankRows += 1
                        continue

                    processedRows += 1

                    logging.debug(member)

                    # remove Unconsumed column names: clientReference, company, policyNumber
                    member.pop("clientReference", None)
                    member.pop("company", None)
                    member.pop("policyNumber", None)

                    # remove any keys with blank in the name
                    for key in list(member.keys()):
                        if "blank" in key:
                            member.pop(key, None)

                    insertFileDataOrg(conn, member)
                    # exit()
                    # break

                wb.close()
                conn.commit()

            except Exception as e:
                logging.error(e)
                conn.rollback()


# insert data into onboardingPolicy table
def insertOnboardingPolicy(conn, data):
    try:
        # check if onboardingPolicy entry exists for policyId
        s = insert(onboardingPolicy).values(data)
        result = conn.execute(s)
        return result.inserted_primary_key[0]
    except Exception as e:
        logging.error(e)
        conn.rollback()

    return None


# update onboardingPolicy status
def updatePolicyStatus(conn, policyId, status, statusNote=None):
    try:
        s = (
            update(onboardingPolicy)
            .where(onboardingPolicy.id == policyId)
            .values(
                status=status,
                statusNote=statusNote,
            )
        )
        conn.execute(s)
    except Exception as e:
        logging.error(e)
        conn.rollback()


# insert data into onboardingData table
def insertOnboardingData(conn, data):
    try:
        s = conn.execute(insert(onboardingData), data)
    except Exception as e:
        logging.error(e)
        conn.rollback()


# @description: get all policies without a brokerageName and schemeName and update the policy with the brokerageName and schemeName from RMA API
# @params: db_conn
def updateBrokerageAndScheme(db_conn):
    try:
        s = (
            onboardingPolicy.__table__.select()
            .with_hint(onboardingPolicy, "WITH (NOLOCK)")
            .where(
                or_(
                    onboardingPolicy.BrokerageName == None,
                    onboardingPolicy.ProviderName == None,
                    onboardingPolicy.ProductOptionId == None,
                )
            )
            .with_only_columns(onboardingPolicy.parentPolicyId)
            .distinct()
            .limit(100)
        )
        result = db_conn.execute(s)
        for row in result.mappings().all():
            logging.debug(f"Check policy: {row}")
            # get brokerageName and schemeName from rma api
            getProvider = utils.get_rma_api(
                f"/clc/api/Policy/Policy/{row["parentPolicyId"]}"
            )
            brokerageName = None
            schemeName = None
            productId = None
            if getProvider:
                brokerageName = getProvider["brokerageName"]
                schemeName = getProvider["clientName"]
                productId = getProvider["productOptionId"]

            # update policy with brokerageName and schemeName
            s = (
                update(onboardingPolicy)
                .where(
                    onboardingPolicy.parentPolicyId == row["parentPolicyId"],
                    or_(
                        onboardingPolicy.BrokerageName == None,
                        onboardingPolicy.ProviderName == None,
                        onboardingPolicy.ProductOptionId == None,
                    ),
                )
                .values(
                    BrokerageName=brokerageName,
                    ProviderName=schemeName,
                    ProductOptionId=productId,
                )
            )
            db_conn.execute(s)
        db_conn.commit()
    except Exception as error:
        logging.error(error)
        db_conn.rollback()


# @description: update beneficiaries and also flag as members
# @params: db_conn
def updateMembersAsBeneficiaries(db_conn):
    s = """with cte as (
select distinct policyId, fileId, fileRow, idNumber 
from onboarding.onboardingData od (nolock)
where memberTypeId = 6
and COALESCE(alsoMember,0) = 0
)
update od
set od.isBeneficiary = 1
from onboarding.onboardingData od 
inner join cte b (nolock) on od.policyId = b.policyId and od.fileId = b.fileId and od.idNumber = b.idNumber
where od.memberTypeId  <> 6
and COALESCE(od.isBeneficiary,0) = 0;

with cte as (
select distinct policyId, fileId, fileRow, idNumber 
from onboarding.onboardingData od (nolock)
where memberTypeId <> 6
and od.isBeneficiary = 1
)
update od
set od.alsoMember = 1, od.exceptions = '[]', od.status = 'New'
from onboarding.onboardingData od 
inner join cte b (nolock) on od.policyId = b.policyId and od.fileId = b.fileId and od.idNumber = b.idNumber
where od.memberTypeId  = 6
and COALESCE(od.alsoMember,0) = 0;

with cte as (
select od.*
from onboarding.onboardingPolicies op (nolock)
inner join onboarding.onboardingData od (nolock) on op.id = od.policyId
where op.status = 'Error' 
)
update op 
set op.status = 'Processing', op.statusNote = 'Checking members'
from onboarding.onboardingPolicies op
where op.status = 'Error'
and not exists(select * from cte where policyId = op.id and status = 'Error');"""
    utils.orm_query(db_conn, s)


# @description: get max cover
# @params: db_conn, parentPolicyId
def getMaxAge(db_conn, parentPolicyId):
    try:
        getProvider = utils.get_rma_api(f"/clc/api/Policy/Policy/{parentPolicyId}")
        # print(getProvider)
        if getProvider:
            productOptionId = getProvider["productOptionId"]
            qry = f"""select pob.productOptionId, max(br.maxAge) as maxAge
    from rules.BenefitRules br (nolock)
    inner join rules.ProductOptionBenefits pob (nolock) on br.benefitId = pob.benefitId and pob.productOptionId = {productOptionId}
    group by pob.productOptionId;"""
            result = utils.orm_select(db_conn, qry)
            # print(result)
            return result[0]["maxAge"]
    except Exception as e:
        logging.error(e)
    return None


# process file to be inserted into onboardingData table
def processFilesNewStructure():
    with utils.orm_conn(os.getenv("DATABASE_URL")) as conn:
        downloaded = controllers.returnFiles(conn)

        for file in downloaded:
            try:
                allMembers = []
                fileId = file["id"]
                # set rows
                totalRows = 0
                blankRows = 0
                processedRows = 0
                hasMainMember = False

                logging.debug(f"Processing {file['fileName']}")
                # open workbook, read only property allows for more consisttent and faster read times esecially when working on larger excel files
                wb = openpyxl.load_workbook(
                    "downloads/" + file["fileName"], read_only=True, data_only=True
                )

                logging.debug(f"Processing {file['fileName']}")

                # set active sheet
                sheet = wb.active

                # get list of column names
                key_names = []
                for x in range(1, sheet.max_column + 1):
                    if sheet.cell(4, x).value:
                        key_names.append(
                            fmt_key(sheet.cell(4, x).value)
                        )  # column names start at row 4

                    # if column is blank then check if rows below are blank then said column name as blank + x
                    else:
                        key_names.append(fmt_key(f"blank{x}"))

                row_values = []
                policyId = None
                spouse = False
                child = False
                extended = False
                getPolicy = None
                spouseCnt = 0
                childCnt = 0

                # get max age
                maxAge = None
                maxAge = getMaxAge(conn, file["providerId"])

                row_num = 5  # rows starts at row 5
                mx_row = sheet.max_row  # last row in sheet
                totalRows = mx_row - row_num + 1

                # convert rows to dictionary
                # iterate over rows
                for row in sheet.iter_rows(min_row=row_num, values_only=True):
                    member = []
                    policyMember = []
                    exceptions = []

                    logging.debug(f"processing {row_num} of {mx_row}")
                    # set dictinary key:value for each column
                    member = {key_names[i]: row[i] for i in range(len(key_names))}
                    hasBlank = True

                    # check if blank row based on these columns being blank
                    for key, value in member.items():
                        if key in [
                            "clientType",
                            "firstName",
                            "surname",
                            "idNumber",
                            "passportNumber",
                        ]:
                            if value:
                                hasBlank = False
                                break

                    # member row
                    member["fileRow"] = row_num
                    row_num += 1
                    if hasBlank:
                        blankRows += 1
                        continue

                    logging.debug(member)
                    processedRows += 1

                    # firstname START
                    # not blank and should be upper case and match to VOPD info
                    member["firstName"] = (
                        str(member["firstName"]).strip().upper().replace(",", "")
                        if member["firstName"]
                        else None
                    )
                    # check if value entered
                    if member["firstName"] == None:
                        exceptions.append(
                            fmt_obj("firstName", "First name cannot be blank")
                        )
                    # check if value contains a number
                    elif utils.contains_number(member["firstName"]):
                        exceptions.append(
                            fmt_obj(
                                "firstName",
                                "First name cannot contain a number",
                            )
                        )
                    # firstname END

                    # surname START
                    # not blank and should be upper case and match to VOPD info
                    member["surname"] = (
                        str(member["surname"]).strip().upper().replace(",", "")
                        if member["surname"]
                        else None
                    )
                    # check if value entered
                    if member["surname"] == None:
                        exceptions.append(fmt_obj("surname", "Surname cannot be blank"))
                    # check if value contains a number
                    elif utils.contains_number(member["surname"]):
                        exceptions.append(
                            fmt_obj("surname", "Surname cannot contain a number")
                        )
                    # surname END

                    # clientType START
                    member["clientType"] = (
                        str(member["clientType"]).lower().strip()
                        if member["clientType"]
                        else None
                    )

                    # standardise clientType
                    member["memberTypeId"] = 0
                    if member["clientType"]:
                        if member["clientType"].lower() in [
                            "extend",
                            "extended",
                            "family member",
                            "familymember",
                        ]:
                            member["clientType"] = "Extended Family"
                        member["clientType"] = string.capwords(member["clientType"])
                        if member["clientType"] not in [
                            "Main Member",
                            "Spouse",
                            "Family Member",
                            "Child",
                            "Extended Family",
                            "Beneficiary",
                        ] and "main" in member["clientType"].lower().replace(" ", ""):
                            member["clientType"] = "Main Member"
                        if member["clientType"] not in [
                            "Main Member",
                            "Spouse",
                            "Family Member",
                            "Child",
                            "Extended Family",
                            "Beneficiary",
                        ]:
                            exceptions.append(
                                fmt_obj(
                                    "MemberType",
                                    "Invalid member type specified.",
                                )
                            )
                            member["memberTypeId"] = 0
                        else:
                            member["memberTypeId"] = utils.returnMemberType(
                                member["clientType"]
                            )
                    elif not member["clientType"]:
                        exceptions.append(
                            fmt_obj(
                                "clientType",
                                "Invalid client type specified.",
                            )
                        )

                    if member["clientType"] == "Spouse":
                        spouse = True
                        spouseCnt += 1
                        if spouseCnt > 1:
                            exceptions.append(
                                fmt_obj(
                                    "clientType",
                                    "Only one spouse is allowed per policy",
                                )
                            )

                    if member["clientType"] == "Child":
                        child = True
                        childCnt += 1
                        if childCnt > 6:
                            exceptions.append(
                                fmt_obj(
                                    "clientType",
                                    "Only six children are allowed per policy",
                                )
                            )
                    if member["clientType"] == "Extended Family":
                        extended = True
                    # clientType END

                    # idNumber START
                    member["idNumber"] = (
                        str(member["idNumber"])
                        .strip()
                        .replace(".", "")
                        .replace("'", "")
                        .replace(" ", "")
                        if member["idNumber"]
                        else None
                    )

                    # add leading zeros to confirm possible ID numbers
                    if (
                        # member["clientType"] == "Child"
                        # and
                        member["idNumber"]
                        and len(member["idNumber"]) in [11, 12]
                    ):
                        member["idNumber"] = ("00" + member["idNumber"])[-13:]

                    if (
                        member["idNumber"] == None
                        and member["clientType"]
                        and "main" in member["clientType"].lower()
                        and member["mainMemberLinkId"]
                    ):
                        member["idNumber"] = member["mainMemberLinkId"]

                    if (
                        member["idNumber"] == None
                        and member["passportNumber"] == None
                        and "main" in member["clientType"].lower()
                        and member["mainMemberLinkId"]
                    ):
                        member["idNumber"] = member["mainMemberLinkId"]

                    passportDate = False

                    # check if valid SA id
                    valid_id = utils.validate_idno(member["idNumber"])
                    member["idTypeId"] = 1
                    if member["idNumber"] == None and member["passportNumber"]:
                        member["idTypeId"] = 2
                        try:
                            if "-" in member["passportNumber"]:
                                member["passportNumber"] = datetime.strptime(
                                    member["passportNumber"][:10], "%Y-%m-%d"
                                )
                        except Exception as e:
                            logging.error(e)

                        member["idNumber"] = member["passportNumber"]
                    elif member["idNumber"] == None and member["dateOfBirth"]:
                        try:
                            member["dateOfBirth"] = datetime.strptime(
                                member["dateOfBirth"], "%Y-%m-%d"
                            )
                            member["idTypeId"] = 2
                            member["idNumber"] = member["dateOfBirth"]
                            passportDate = True
                        except Exception as e:
                            logging.error(e)

                    # member["idTypeId"] = 1 if valid_id else 2
                    if (
                        member["idTypeId"] == 1
                        and not valid_id
                        and member["clientType"]
                        and "main" not in member["clientType"].lower()
                        and not member["dateOfBirth"]
                    ):
                        try:
                            check_date = datetime.strptime(
                                member["idNumber"][:6], "%y%m%d"
                            )
                            member["dateOfBirth"] = check_date
                            member["idNumber"] = member["dateOfBirth"]
                            member["idTypeId"] = 2
                        except:
                            member["dateOfBirth"] = None

                    if member["idTypeId"] == 1 and not valid_id:
                        exceptions.append(
                            fmt_obj(
                                "idNumber",
                                "Invalid ID number supplied",
                            )
                        )

                    if (
                        not (member["idNumber"])
                        and not (member["passportNumber"])
                        and member["clientType"]
                        in [
                            "Main Member",
                            # "Spouse",
                        ]
                    ):
                        exceptions.append(
                            fmt_obj(
                                "idNumber",
                                "No ID number or passport supplied",
                            )
                        )

                    # idNumber END

                    # dob check START
                    # if valid id calculate DOB
                    member["dateOfBirth"] = (
                        utils.return_dob(member["idNumber"])
                        if valid_id
                        else member["dateOfBirth"]
                    )

                    try:

                        if type(member["dateOfBirth"]) == str:
                            check_date = datetime.strptime(
                                member["dateOfBirth"], "%Y-%m-%d"
                            )
                        else:
                            check_date = member["dateOfBirth"]
                        # member["dateOfBirth"] = check_date.strftime("%Y-%m-%d")
                        member["dateOfBirth"] = check_date
                    except:
                        member["dateOfBirth"] = None

                    if not (member["dateOfBirth"]):
                        try:
                            if member["idNumber"] and "-" in member["idNumber"]:
                                check_date = datetime.strptime(
                                    member["idNumber"], "%d-%m-%y"
                                )
                            elif member["idNumber"] and "/" in member["idNumber"]:
                                check_date = datetime.strptime(
                                    member["idNumber"], "%d/%m/%y"
                                )
                            elif (
                                not (member["idNumber"])
                                and member["passportNumber"]
                                and "-" in member["passportNumber"]
                            ):
                                # check_date = datetime.strptime(
                                #     member["passportNumber"], "%d-%m-%Y"
                                # )
                                check_date = member["passportNumber"]
                                passportDate = True
                            elif (
                                not (member["idNumber"])
                                and member["passportNumber"]
                                and "/" in member["passportNumber"]
                            ):
                                check_date = datetime.strptime(
                                    member["passportNumber"], "%d/%m/%Y"
                                )
                                passportDate = True
                            elif not (member["idNumber"]) and member["passportNumber"]:
                                # check_date = datetime.strptime(
                                #     member["passportNumber"], "%Y-%m-%d %H:%M:%S"
                                # )
                                check_date = member["passportNumber"]
                                passportDate = True
                            else:
                                check_date = datetime.strptime(
                                    member["idNumber"][:6], "%y%m%d"
                                )

                            # member["dateOfBirth"] = check_date.strftime("%Y-%m-%d")
                            member["dateOfBirth"] = check_date

                        except:
                            member["dateOfBirth"] = None

                    if not (member["dateOfBirth"]):
                        try:
                            if member["idNumber"] and "-" in member["idNumber"]:
                                check_date = datetime.strptime(
                                    member["idNumber"], "%Y-%m-%d"
                                )
                            elif member["idNumber"] and "/" in member["idNumber"]:
                                check_date = datetime.strptime(
                                    member["idNumber"], "%Y/%m/%d"
                                )
                            elif (
                                not (member["idNumber"])
                                and member["passportNumber"]
                                and "-" in member["passportNumber"]
                            ):
                                check_date = datetime.strptime(
                                    member["passportNumber"], "%Y-%m-%d"
                                )
                                passportDate = True

                            elif (
                                not (member["idNumber"])
                                and member["passportNumber"]
                                and "/" in member["passportNumber"]
                            ):
                                check_date = datetime.strptime(
                                    member["passportNumber"], "%Y/%m/%d"
                                )
                                passportDate = True

                            member["dateOfBirth"] = check_date
                        except:
                            member["dateOfBirth"] = None

                    if not (member["dateOfBirth"]) and isinstance(
                        member["idNumber"], datetime
                    ):
                        member["dateOfBirth"] = member["idNumber"]

                    # check

                    try:

                        if (
                            member["idTypeId"] == 2
                            and member["dateOfBirth"]
                            and not passportDate
                            and isinstance(member["idNumber"], datetime)
                            and member["idNumber"] == member["dateOfBirth"]
                        ):
                            passportDate = True

                    except Exception as e:
                        logging.error(e)

                    try:

                        if (
                            member["idTypeId"] == 2
                            and member["dateOfBirth"]
                            and not passportDate
                        ):
                            # check if first six 10 digits match date of birth set to yyyy-mm-dd
                            if member["idNumber"][:6] == member["dateOfBirth"].strftime(
                                "%y%m%d"
                            ):
                                passportDate = True

                    except Exception as e:
                        logging.error(e)

                    # if member["idTypeId"] == 2:
                    #     print(passportDate)
                    #     print(member)
                    #     conn.rollback()
                    #     exit()

                    member["passportNumber"] = (
                        str(member["passportNumber"]).strip().replace(" ", "")
                        if "passportNumber" in member
                        and member["passportNumber"]
                        and not passportDate
                        else None
                    )

                    # correction for date of birth being in the future
                    try:

                        if (
                            member["dateOfBirth"]
                            and member["dateOfBirth"] > datetime.now()
                        ):
                            member["dateOfBirth"] = member[
                                "dateOfBirth"
                            ] - relativedelta(years=100)
                    except Exception as e:
                        logging.error("date of birth is not a date")
                        member["dateOfBirth"] = None

                    if not (member["dateOfBirth"]):

                        exceptions.append(
                            fmt_obj(
                                "dateOfBirth",
                                "No valid DOB supplied",
                            )
                        )
                    # dob check END

                    # check age
                    if member["dateOfBirth"]:

                        age = utils.return_age(
                            member["dateOfBirth"],
                            file["joinDate"],
                        )

                        if age and member["clientType"] == "Main Member" and age < 18:
                            exceptions.append(
                                fmt_obj(
                                    "dateOfBirth",
                                    "Main member must be 18 years or older",
                                )
                            )

                        if age and member["clientType"] == "Beneficiary" and age < 18:
                            exceptions.append(
                                fmt_obj(
                                    "dateOfBirth",
                                    "Beneficiary must be 18 years or older",
                                )
                            )

                        # check dob child
                        if age and member["clientType"] == "Child" and age > 21:
                            tempDOB = member["dateOfBirth"] + relativedelta(years=100)

                            age2 = utils.return_age(
                                tempDOB,
                                file["joinDate"],
                            )

                            if age2 >= 0:
                                age = age2

                            if age > 25:
                                exceptions.append(
                                    fmt_obj(
                                        "dateOfBirth",
                                        "Child must be younger than 25 years",
                                    )
                                )
                            else:
                                member["dateOfBirth"] = tempDOB.strftime("%Y-%m-%d")

                        if age and member["clientType"] == "Main Member":

                            # print(f"Max age is {maxAge}")
                            if maxAge and age > maxAge:
                                exceptions.append(
                                    fmt_obj(
                                        "dateOfBirth",
                                        f"Main member exceeds maximum age of {maxAge} years",
                                    )
                                )
                                # print(f"Main member exceeds maximum age of {maxAge} years and age is {age}")
                                # conn.rollback()
                                # exit()

                        if age and age < 0:
                            exceptions.append(
                                fmt_obj(
                                    "dateOfBirth",
                                    "Invalid date of birth supplied",
                                )
                            )

                    # check VOPD for ID number START
                    # if valid id check VOPD
                    member["status"] = "New"
                    member["idValid"] = True if valid_id else False
                    if valid_id:

                        # check for similary update date of birth and name
                        getMember = controllers.returnVOPD(conn, member["idNumber"])

                        if getMember:
                            # update name and surname if similarity is greater than 50%
                            firstNameSimilarity = utils.similarityScores(
                                getMember[0]["firstName"], member["firstName"]
                            )
                            if (
                                firstNameSimilarity["percentage"] > 0
                                or firstNameSimilarity["levenshtein"] >= 0.5
                            ):
                                member["firstName"] = getMember[0]["firstName"]
                            else:
                                exceptions.append(
                                    fmt_obj(
                                        "firstName",
                                        f"First name does not match VOPD - {getMember[0]['firstName']}",
                                    )
                                )

                            surnameSimilarity = utils.similarityScores(
                                getMember[0]["surname"], member["surname"]
                            )

                            if (
                                surnameSimilarity["percentage"] > 0
                                or surnameSimilarity["levenshtein"] >= 0.5
                            ):
                                member["surname"] = getMember[0]["surname"]
                            else:
                                exceptions.append(
                                    fmt_obj(
                                        "surname",
                                        f"Surname does not match VOPD {getMember[0]['surname']}",
                                    )
                                )

                            member["dateOfBirth"] = getMember[0]["dateOfBirth"]
                            member["dateOfDeath"] = None
                            # if person deceased then rerunning vopd doesn't matter
                            if getMember[0]["dateOfDeath"]:
                                member["dateOfDeath"] = getMember[0]["dateOfDeath"]
                                exceptions.append(
                                    fmt_obj(
                                        "idNumber",
                                        "Member is deceased",
                                    )
                                )
                                member["vopdVerified"] = True
                                member["vopdVerificationDate"] = getMember[0][
                                    "updatedAt"
                                ]
                                member["status"] = "Error"
                            # if checked in the last 30 days
                            elif getMember[0]["updatedAt"] >= (
                                datetime.now() - timedelta(days=30)
                            ):
                                member["vopdVerified"] = True
                                member["vopdVerificationDate"] = getMember[0][
                                    "updatedAt"
                                ]
                                member["status"] = "VOPD Complete"
                        # elif member["clientType"] == "Main Member":
                        else:
                            if controllers.addVOPD(conn, member["idNumber"]):
                                member["status"] = "Waiting for VOPD"
                    # check VOPD for ID number END

                    # flag out support documents required

                    if member["idTypeId"] == 2:
                        if member["clientType"] == "Main Member":
                            exceptions.append(
                                {
                                    "field": "idTypeId",
                                    "message": "Main member ID type is passport and no supporting document",
                                }
                            )
                        # check if member["passportNumber"] is a date
                        elif not passportDate:
                            exceptions.append(
                                {
                                    "field": "idTypeId",
                                    "message": "Member ID type is passport and no supporting document",
                                }
                            )

                    # email check START
                    member["email"] = (
                        str(member["email"]).strip() if member["email"] else None
                    )
                    if member["email"] and not (utils.valid_email(member["email"])):
                        exceptions.append(
                            fmt_obj(
                                "email",
                                "Invalid e-mail address supplied, check format.",
                            )
                        )

                    # if main member and no e-mail then flag as error
                    # if not member["email"] and member["clientType"] == "Main Member":
                    #     exceptions.append(
                    #         fmt_obj(
                    #             "email",
                    #             "No e-mail address supplied for main member",
                    #         )
                    #     )
                    # email check END

                    # if telephone does not start with 0 and is 9 digits long then add 0 to start
                    member["mobile"] = (
                        str(member["mobile"]).strip() if member["mobile"] else None
                    )

                    # if main member and no mobile then flag as error
                    if not member["mobile"] and member["clientType"] == "Main Member":
                        exceptions.append(
                            fmt_obj(
                                "mobile",
                                "No mobile number supplied for main member",
                            )
                        )

                    if member["mobile"] and len(str(member["mobile"])) == 9:
                        if not str(member["mobile"]).startswith("0"):
                            member["mobile"] = "0" + str(member["mobile"])

                    # if mobile is specified and 11 digits long and start with 27
                    if (
                        member["mobile"]
                        and len(str(member["mobile"])) == 11
                        and str(member["mobile"]).startswith("27")
                    ):
                        member["mobile"] = "0" + str(member["mobile"])[2:]

                    if member["mobile"] and not utils.valid_phone(
                        str(member["mobile"]), True
                    ):
                        exceptions.append(
                            fmt_obj(
                                "mobile",
                                "Invalid Contact Numbers",
                            )
                        )

                    # preferredMethodOfCommunication check START
                    member["preferredMethodOfCommunication"] = (
                        str(member["preferredMethodOfCommunication"])
                        .strip()
                        .replace("-", "")
                        if member["preferredMethodOfCommunication"]
                        else None
                    )

                    if member["preferredMethodOfCommunication"]:
                        if (
                            str(member["preferredMethodOfCommunication"]).lower()
                            == "email"
                        ):
                            member["preferredMethodOfCommunication"] = "Email"

                        elif (
                            str(member["preferredMethodOfCommunication"]).lower()
                            == "sms"
                        ):
                            member["preferredMethodOfCommunication"] = "SMS"

                        else:
                            exceptions.append(
                                fmt_obj(
                                    "preferredMethodOfCommunication",
                                    "Invalid preference of communication. Only Email and SMS are allowed.",
                                )
                            )
                    if not member["preferredMethodOfCommunication"]:
                        member["preferredMethodOfCommunication"] = "Email"

                    # Map to IDs
                    if member["preferredMethodOfCommunication"] == "Email":
                        member["preferredMethodOfCommunication"] = 1
                    elif member["preferredMethodOfCommunication"] == "SMS":
                        member["preferredMethodOfCommunication"] = 3
                    else:
                        member["preferredMethodOfCommunication"] = 1

                    if (
                        not member["preferredMethodOfCommunication"]
                        and member["clientType"] == "Main Member"
                    ):
                        exceptions.append(
                            fmt_obj(
                                "preferredMethodOfCommunication",
                                "No preferred method of communication specified",
                            )
                        )

                    # if preferredMethodOfCommunication == 1 and not email then error
                    if (
                        member["preferredMethodOfCommunication"] == 1
                        and not member["email"]
                    ):
                        exceptions.append(
                            fmt_obj(
                                "preferredMethodOfCommunication",
                                "Email is required for preferred method of communication",
                            )
                        )

                    # if preferredMethodOfCommunication == 3 and not mobile then error
                    if (
                        member["preferredMethodOfCommunication"] == 3
                        and not member["mobile"]
                    ):
                        exceptions.append(
                            fmt_obj(
                                "preferredMethodOfCommunication",
                                "Mobile number is required for preferred method of communication",
                            )
                        )

                    # preferredMethodOfCommunication check END

                    member["address1"] = (
                        str(member["address1"]).upper().strip()
                        if member["address1"]
                        else None
                    )
                    member["address2"] = (
                        str(member["address2"]).upper().strip()
                        if member["address2"]
                        else None
                    )
                    member["city"] = (
                        str(member["city"]).upper().strip() if member["city"] else None
                    )
                    member["province"] = (
                        str(member["province"]).upper().strip()
                        if member["province"]
                        else None
                    )
                    member["country"] = (
                        str(member["country"]).upper().strip()
                        if member["country"]
                        else None
                    )

                    member["areaCode"] = (
                        str(member["areaCode"]).strip() if member["areaCode"] else None
                    )

                    if (
                        member["address1"]
                        and member["address2"]
                        and member["address2"] == member["address1"]
                    ):
                        # remove address2 from address1 and trim
                        member["address2"] = None

                    if (
                        member["address1"]
                        and member["address2"]
                        and member["address2"] in member["address1"]
                    ):
                        # remove address2 from address1 and trim
                        member["address1"] = (
                            member["address1"].replace(member["address2"], "").strip()
                        )

                    if (
                        member["address1"]
                        and member["city"]
                        and member["city"] == member["address1"]
                    ):
                        # remove address2 from address1 and trim
                        member["city"] = None

                    # remove city from address1 and trim
                    if (
                        member["address1"]
                        and member["city"]
                        and member["city"]
                        and member["city"] in member["address1"]
                    ):
                        member["address1"] = (
                            member["address1"].replace(member["city"], "").strip()
                        )

                    if (
                        member["address1"]
                        and member["province"]
                        and member["province"] == member["address1"]
                    ):
                        # remove address2 from address1 and trim
                        member["province"] = None

                    # if province and address 1 then check if province is in address 1 and remove and trim
                    if (
                        member["province"]
                        and member["address1"]
                        and member["province"] in member["address1"]
                    ):
                        member["address1"] = (
                            member["address1"].replace(member["province"], "").strip()
                        )

                    if (
                        member["address1"]
                        and member["areaCode"]
                        and member["areaCode"] == member["address1"]
                    ):
                        # remove address2 from address1 and trim
                        member["areaCode"] = None

                    # if postal code and address 1 then check if postal code is in address 1 and remove and trim
                    if (
                        member["areaCode"]
                        and member["address1"]
                        and member["areaCode"] in member["address1"]
                    ):
                        member["address1"] = (
                            member["address1"].replace(member["areaCode"], "").strip()
                        )

                    if not member["address1"] and member["clientType"] == "Main Member":
                        exceptions.append(
                            fmt_obj(
                                "address1",
                                "No address details specified",
                            )
                        )

                    # if member["address1"] and address1 is greater than 50 characters then error
                    if member["address1"] and len(member["address1"]) > 50:
                        exceptions.append(
                            fmt_obj(
                                "address1",
                                "Address 1 cannot be longer than 50 characters",
                            )
                        )

                    # if member["address2"] and address1 is greater than 50 characters then error
                    if member["address2"] and len(member["address2"]) > 50:
                        exceptions.append(
                            fmt_obj(
                                "address2",
                                "Address 2 cannot be longer than 50 characters",
                            )
                        )

                    # remove all non numbers from areaCode
                    if member["areaCode"]:
                        member["areaCode"] = re.sub(r"[^0-9]", "", member["areaCode"])

                    # if areaCode is specified and longer than 6 characters then clear
                    if member["areaCode"] and len(member["areaCode"]) > 6:
                        member["areaCode"] = None

                    # if areaCode is specified and longer than 4 characters then error
                    if member["areaCode"] and len(member["areaCode"]) > 4:
                        exceptions.append(
                            fmt_obj(
                                "areaCode",
                                "Area code cannot be longer than 4 characters",
                            )
                        )

                    list_of_provinces = [
                        "EASTERN CAPE",
                        "FREE STATE",
                        "GAUTENG",
                        "KWAZULU-NATAL",
                        "LIMPOPO",
                        "MPUMALANGA",
                        "NORTH WEST",
                        "NORTHERN CAPE",
                        "WESTERN CAPE",
                    ]

                    # check for variation of provinces specified and correct
                    province_patterns = {
                        "EASTERN CAPE": r"EAST.*CAPE",
                        "GAUTENG": r"(joha.*|gauteng)",
                        "KWAZULU-NATAL": r"KWA.*ZU.*",
                        "NORTHERN CAPE": r"NORT.*CAP.*",
                        "LIMPOPO": r"LIMP.*",
                        "NORTH WEST": r"NORT.*W.*",
                        "MPUMALANGA": r"MPUM.*",
                        "FREE STATE": r"FREEST.*",
                    }

                    if "province" in member and member["province"]:
                        province = member["province"].strip().upper()

                        # Try to update province based on matching pattern
                        for correct_province, pattern in province_patterns.items():
                            if re.match(pattern, province):
                                member["province"] = correct_province
                                break

                    if (
                        "province" in member
                        and member["province"]
                        and member["province"] not in list_of_provinces
                    ):
                        exceptions.append(
                            fmt_obj(
                                "province",
                                f"Invalid province specified - {member["province"]}",
                            )
                        )
                        member["province"] = None

                    member["postalAddress1"] = (
                        str(member["postalAddress1"]).upper().strip()
                        if member["postalAddress1"]
                        else None
                    )
                    member["postalAddress2"] = (
                        str(member["postalAddress2"]).upper().strip()
                        if member["postalAddress2"]
                        else None
                    )
                    member["postalCity"] = (
                        str(member["postalCity"]).upper().strip()
                        if member["postalCity"]
                        else None
                    )
                    member["postalProvince"] = (
                        str(member["postalProvince"]).upper().strip()
                        if member["postalProvince"]
                        else None
                    )
                    member["postalCountry"] = (
                        str(member["postalCountry"]).upper().strip()
                        if member["postalCountry"]
                        else None
                    )

                    member["postalCode"] = (
                        str(member["postalCode"]).strip()
                        if member["postalCode"]
                        else None
                    )

                    if (
                        member["postalAddress1"]
                        and member["postalAddress2"]
                        and member["postalAddress2"] in member["postalAddress1"]
                    ):
                        # remove address2 from address1 and trim
                        member["postalAddress1"] = (
                            member["postalAddress1"]
                            .replace(member["postalAddress2"], "")
                            .strip()
                        )

                    # remove city from address1 and trim
                    if (
                        member["postalAddress1"]
                        and member["postalCity"]
                        and member["postalCity"]
                        and member["postalCity"] in member["postalAddress1"]
                    ):
                        member["postalAddress1"] = (
                            member["postalAddress1"]
                            .replace(member["postalCity"], "")
                            .strip()
                        )

                    # if province and address 1 then check if province is in address 1 and remove and trim
                    if (
                        member["postalProvince"]
                        and member["postalAddress1"]
                        and member["postalProvince"] in member["postalAddress1"]
                    ):
                        member["postalAddress1"] = (
                            member["postalAddress1"]
                            .replace(member["postalProvince"], "")
                            .strip()
                        )

                    # if postal code and address 1 then check if postal code is in address 1 and remove and trim

                    if (
                        member["postalCode"]
                        and member["postalAddress1"]
                        and str(member["postalCode"]) in member["postalAddress1"]
                    ):
                        member["postalAddress1"] = (
                            member["postalAddress1"]
                            .replace(str(member["postalCode"]), "")
                            .strip()
                        )

                    # previous insurance check START

                    member["previousInsurerPolicyNumber"] = (
                        str(member["previousInsurerPolicyNumber"]).upper().strip()
                        if member["previousInsurerPolicyNumber"]
                        else None
                    )

                    member["previousInsurerJoinDate"] = (
                        member["previousInsurerJoinDate"]
                        if member["previousInsurerJoinDate"]
                        and type(member["previousInsurerJoinDate"]) == datetime
                        else None
                    )

                    member["previousInsurerCancellationDate"] = (
                        member["previousInsurerCancellationDate"]
                        if member["previousInsurerCancellationDate"]
                        and type(member["previousInsurerJoinDate"]) == datetime
                        else None
                    )

                    member["PreviousInsurerCoverAmount"] = 0

                    if member["clientType"] == "Main Member" or not policyId:
                        if member["clientType"] == "Main Member":
                            hasMainMember = True
                            policyId = None

                        logging.debug(member["benefitName"])
                        coverAmount = utils.returnCoverAmount(member["benefitName"])
                        if member["clientType"] != "Beneficiary":
                            # get cover amount from policy
                            if coverAmount == 0:
                                benefits = controllers.getMemberBenefitByName(
                                    conn, member["benefitName"]
                                )
                                if benefits and len(benefits) > 0:
                                    coverAmount = benefits[0]["benefitAmount"]
                                elif benefits:
                                    coverAmount = benefits["benefitAmount"]

                        policyStatus = "Processing"
                        statusNote = "Checking members"

                        # if coverAmount == 0 and not benefitExists:
                        #     coverAmount = 0

                        # if coverAmount == 0:
                        #     policyStatus = "Error"
                        #     statusNote = "Could not allocate cover amount to policy"

                        if coverAmount == 0 or not coverAmount:

                            # add exception
                            exceptions.append(
                                fmt_obj(
                                    "benefitName",
                                    "Benefit name not found",
                                )
                            )
                            policyStatus = "Error"
                            statusNote = "Could not allocate cover amount to policy"

                        spouse = False
                        child = False
                        extended = False
                        childCnt = 0
                        spouseCnt = 0

                        if not policyId:
                            policyData = {
                                "fileId": str(fileId).lower(),
                                "PolicyInceptionDate": file["joinDate"],
                                "parentPolicyId": file["providerId"],
                                "createdBy": file["createdBy"],
                                "status": policyStatus,
                                "statusNote": statusNote,
                                "coverAmount": coverAmount,
                            }
                            policyId = insertOnboardingPolicy(conn, policyData)
                            logging.debug(f"Policy created {policyId}")

                    member["policyId"] = policyId
                    member["coverAmount"] = coverAmount
                    member["memberType"] = member["clientType"]
                    member["fileId"] = str(fileId).lower()
                    member["isBeneficiary"] = (
                        True if member["memberTypeId"] in [1, 6] else False
                    )
                    member["joinDate"] = file["joinDate"]

                    if len(exceptions) > 0:
                        policyStatus = "Error"
                        statusNote = "Issue with members"
                        member["status"] = "Error"

                        updatePolicyStatus(conn, policyId, policyStatus, statusNote)

                        # update policy status

                    member.pop("clientReference", None)
                    member.pop("company", None)
                    member.pop("policyNumber", None)
                    member.pop("passportNumber", None)
                    member.pop("clientType", None)

                    # set member exceptions
                    member["exceptions"] = json.dumps(exceptions)

                    # remove any keys with blank in the name
                    for key in list(member.keys()):
                        if "blank" in key:
                            member.pop(key, None)

                    logging.debug(member)

                    allMembers.append(member)

                    # exit()

                    if len(allMembers) % 2000 == 0:
                        insertOnboardingData(conn, allMembers)
                        allMembers = []

                logging.debug(f"hasMainMember {hasMainMember}")

                if hasMainMember:
                    if len(allMembers) > 0:
                        insertOnboardingData(conn, allMembers)
                        allMembers = []
                    conn.commit()
                    updateMembersAsBeneficiaries(conn)
                    updateBrokerageAndScheme(conn)

                    controllers.updateFileStatus(
                        conn,
                        file["id"],
                        "Uploaded",
                        "Uploaded members",
                        totalRows,
                        blankRows,
                        processedRows,
                    )

                    # utils.sendEmailNotification(
                    #     file["createdBy"],
                    #     "RMA Client Connect - File Processed",
                    #     f"Your file, {file["orgFileName"]} has been loaded successfully and is being processed.",
                    #     contentType="HTML",
                    # )
                    controllers.insertNotification(
                        conn,
                        file["createdBy"],
                        "File Processed",
                        f"Your file, {file["orgFileName"]} has been loaded successfully and is being processed.",
                        link=f"{os.getenv("APP_LINK")}/MyPolicies?fileId={file["id"]}",
                    )
                else:
                    logging.error("No main member")
                    conn.rollback()
                    controllers.updateFileStatus(
                        conn,
                        file["id"],
                        "Error",
                        "Please check file - Member type issue",
                        totalRows,
                    )
                    utils.sendEmailNotification(
                        file["createdBy"],
                        "RMA Client Connect - File Processed",
                        f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                    )
                    controllers.insertNotification(
                        conn,
                        file["createdBy"],
                        "File Processing Error",
                        f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                        "error",
                    )

                wb.close()
            except Exception as e:
                logging.error(f"error: {e}")
                # logging.error("Traceback: %s", traceback.format_exc())
                conn.rollback()
                controllers.updateFileStatus(
                    conn,
                    file["id"],
                    "Error",
                    "Please check file - Member type issue",
                    totalRows,
                )
                utils.sendEmailNotification(
                    file["createdBy"],
                    "RMA Client Connect - File Processed",
                    f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                )
                controllers.insertNotification(
                    conn,
                    file["createdBy"],
                    "File Processing Error",
                    f"Your file, {file["orgFileName"]} has failed to load. Please check the file and try again",
                    "error",
                )
